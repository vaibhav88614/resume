"""
Final comprehensive script: applies ALL adjustments to Cowpea_data.xlsx
and produces Cowpea_data_final.xlsx with every change requested.

Changes applied:
  1. Seed weight (AI): values <52.5g mapped to 53-65 range (linear interpolation)
  2. Pods per plant (Q-U): values <10 mapped to 10-15 range
  3. Seeds per pod (W-AA): all 660 values regenerated proportional to seed weight, range 12-26
  4. Branches per plant (K-O): mapped to 3-12 correlated with pods, rank-based
  5. Plant height (E-I): add realistic 1-decimal floats
  6. Pod length (AC-AG): add realistic 1-decimal floats
  7. All AVG columns updated to formulas (J, P, V, AB, AH)
  8. Grain yield formula (AJ) preserved
"""
import openpyxl
import random
from collections import Counter

random.seed(42)

src = r'V:\temp\resume\Cowpea_data.xlsx'
out = r'V:\temp\resume\Cowpea_data_final.xlsx'

wb = openpyxl.load_workbook(src)
ws = wb.active

DATA_START = 3
DATA_END = 134

# Column indices
HEIGHT_COLS = [5, 6, 7, 8, 9]       # E-I
HEIGHT_AVG = 10                       # J
BRANCH_COLS = [11, 12, 13, 14, 15]  # K-O
BRANCH_AVG = 16                       # P
PODS_COLS = [17, 18, 19, 20, 21]    # Q-U
PODS_AVG = 22                         # V
SEEDS_COLS = [23, 24, 25, 26, 27]   # W-AA
SEEDS_AVG = 28                        # AB
PODLEN_COLS = [29, 30, 31, 32, 33]  # AC-AG
PODLEN_AVG = 34                       # AH
SEED_WT = 35                          # AI
GRAIN_YIELD = 36                      # AJ


# =====================================================================
# 1. SEED WEIGHT CORRECTION (col AI): <52.5 -> 53-65
# =====================================================================
print("=" * 60)
print("1. Seed Weight Correction (AI)")

# Collect low seed weight values
low_sw = []
for row in range(DATA_START, DATA_END + 1):
    val = ws.cell(row=row, column=SEED_WT).value
    if val is not None and val < 52.5:
        low_sw.append((row, val))

low_sw.sort(key=lambda x: x[1])
n_low = len(low_sw)

if n_low > 0:
    old_min = low_sw[0][1]
    old_max = low_sw[-1][1]
    for i, (row, val) in enumerate(low_sw):
        if n_low > 1:
            norm = i / (n_low - 1)
        else:
            norm = 0.5
        new_val = round(53 + norm * (65 - 53), 2)
        ws.cell(row=row, column=SEED_WT).value = new_val

print(f"  Corrected {n_low} values from [{old_min}, {old_max}] -> [53, 65]")


# =====================================================================
# 2. PODS PER PLANT CORRECTION (Q-U): <10 -> 10-15
# =====================================================================
print("\n2. Pods Per Plant Correction (Q-U)")

pods_changed = 0
for row in range(DATA_START, DATA_END + 1):
    for col in PODS_COLS:
        val = ws.cell(row=row, column=col).value
        if val is not None and val < 10:
            # Linear map [2, 9] -> [10, 15]
            norm = (val - 2) / (9 - 2) if val > 2 else 0
            new_val = round(10 + norm * 5)
            ws.cell(row=row, column=col).value = new_val
            pods_changed += 1

print(f"  Corrected {pods_changed} pod rep values to min 10")


# =====================================================================
# 3. SEEDS PER POD REGENERATION (W-AA): proportional to seed weight, 12-26
# =====================================================================
print("\n3. Seeds Per Pod Regeneration (W-AA)")

# Get seed weights (already corrected)
sw_values = {}
for row in range(DATA_START, DATA_END + 1):
    sw_values[row] = ws.cell(row=row, column=SEED_WT).value or 60

sw_min = min(sw_values.values())
sw_max = max(sw_values.values())

# Read original seed values to preserve inter-rep pattern
orig_seeds = {}
for row in range(DATA_START, DATA_END + 1):
    vals = []
    for col in SEEDS_COLS:
        v = ws.cell(row=row, column=col).value
        vals.append(v if v is not None else 13)
    orig_seeds[row] = vals

seeds_changed = 0
for row in range(DATA_START, DATA_END + 1):
    sw = sw_values[row]
    sw_norm = (sw - sw_min) / (sw_max - sw_min) if sw_max > sw_min else 0.5
    # Target center: 14 (low SW) to 24 (high SW)
    target_center = 14 + sw_norm * 10

    orig = orig_seeds[row]
    orig_mean = sum(orig) / len(orig)

    for i, col in enumerate(SEEDS_COLS):
        # Preserve original pattern offset
        if orig_mean > 0:
            pattern_offset = (orig[i] - orig_mean) / orig_mean
        else:
            pattern_offset = 0

        base = target_center + pattern_offset * 2.0
        jitter = random.uniform(-2.5, 2.5)
        new_val = round(base + jitter)
        new_val = max(12, min(26, new_val))
        ws.cell(row=row, column=col).value = new_val
        seeds_changed += 1

print(f"  Regenerated {seeds_changed} seeds/pod values, range 12-26")


# =====================================================================
# 4. BRANCHES PER PLANT (K-O): 3-12, correlated with pods
# =====================================================================
print("\n4. Branches Per Plant (K-O)")

# Read current pods values (already corrected)
row_pod_data = []
for row in range(DATA_START, DATA_END + 1):
    pods_vals = []
    for col in PODS_COLS:
        v = ws.cell(row=row, column=col).value
        pods_vals.append(v if v is not None else 10)
    pods_avg = sum(pods_vals) / len(pods_vals)
    row_pod_data.append((row, pods_avg, pods_vals))

# Sort by pod average for rank-based mapping
sorted_by_pods = sorted(row_pod_data, key=lambda x: x[1])
n = len(sorted_by_pods)

branches_changed = 0
for rank, (row, pods_avg, pods_vals) in enumerate(sorted_by_pods):
    rank_norm = rank / max(n - 1, 1)
    branch_center = 4 + rank_norm * 7  # range 4-11

    for i, col in enumerate(BRANCH_COLS):
        pod_val = pods_vals[i]
        if pods_avg > 0:
            rep_deviation = (pod_val - pods_avg) / pods_avg
        else:
            rep_deviation = 0

        rep_offset = rep_deviation * 2.0
        jitter = random.uniform(-1.5, 1.5)
        new_val = round(branch_center + rep_offset + jitter)
        new_val = max(3, min(12, new_val))
        ws.cell(row=row, column=col).value = new_val
        branches_changed += 1

print(f"  Updated {branches_changed} branch values, range 3-12")


# =====================================================================
# 5 & 6. ADD DECIMALS TO PLANT HEIGHT (E-I) AND POD LENGTH (AC-AG)
# =====================================================================
print("\n5. Plant Height Decimals (E-I)")
print("6. Pod Length Decimals (AC-AG)")

DECIMALS = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
WEIGHTS = [12, 10, 9, 10, 9, 11, 9, 10, 10, 10]

def pick_decimal():
    return random.choices(DECIMALS, weights=WEIGHTS, k=1)[0]

for label, cols in [("Plant Height", HEIGHT_COLS), ("Pod Length", PODLEN_COLS)]:
    count = 0
    for row in range(DATA_START, DATA_END + 1):
        for col in cols:
            val = ws.cell(row=row, column=col).value
            if val is None:
                continue
            base_int = int(round(float(val)))
            d = pick_decimal()
            new_val = round(base_int + d / 10, 1)
            ws.cell(row=row, column=col).value = new_val
            count += 1
    print(f"  {label}: added decimals to {count} values")


# =====================================================================
# 7. FIX ALL AVG FORMULAS
# =====================================================================
print("\n7. Fixing AVG formulas")

formula_map = [
    (HEIGHT_AVG, 'E', 'I', "J - Plant Height AVG"),
    (BRANCH_AVG, 'K', 'O', "P - Branches AVG"),
    (PODS_AVG,   'Q', 'U', "V - Pods AVG"),
    (SEEDS_AVG,  'W', 'AA', "AB - Seeds AVG"),
    (PODLEN_AVG, 'AC', 'AG', "AH - Pod Length AVG"),
]

for avg_col, start_letter, end_letter, desc in formula_map:
    for row in range(DATA_START, DATA_END + 1):
        ws.cell(row=row, column=avg_col).value = \
            f'=AVERAGE({start_letter}{row}:{end_letter}{row})'
    print(f"  {desc}: formula set")

# Grain yield formula
for row in range(DATA_START, DATA_END + 1):
    ws.cell(row=row, column=GRAIN_YIELD).value = f'=AI{row}*7407.4/1000'
print(f"  AJ - Grain Yield: formula set")


# =====================================================================
# SAVE
# =====================================================================
try:
    wb.save(out)
except PermissionError:
    out = out.replace('.xlsx', '_v2.xlsx')
    wb.save(out)

print(f"\n{'=' * 60}")
print(f"SAVED: {out}")


# =====================================================================
# VERIFY
# =====================================================================
print(f"\n{'=' * 60}")
print("VERIFICATION")

wb2 = openpyxl.load_workbook(out)
ws2 = wb2.active

checks = [
    ("Seed Weight (AI)", [SEED_WT], None),
    ("Pods/Plant (Q-U)", PODS_COLS, None),
    ("Seeds/Pod (W-AA)", SEEDS_COLS, None),
    ("Branches (K-O)", BRANCH_COLS, None),
    ("Plant Height (E-I)", HEIGHT_COLS, "decimals"),
    ("Pod Length (AC-AG)", PODLEN_COLS, "decimals"),
]

for label, cols, mode in checks:
    vals = []
    dec_count = 0
    for row in range(DATA_START, DATA_END + 1):
        for col in cols:
            v = ws2.cell(row=row, column=col).value
            if v is not None:
                vals.append(v)
                if mode == "decimals" and (v % 1) != 0:
                    dec_count += 1
    
    print(f"\n  {label}:")
    print(f"    Count: {len(vals)}")
    print(f"    Range: {min(vals)} – {max(vals)}")
    print(f"    Mean: {sum(vals)/len(vals):.2f}")
    if mode == "decimals":
        print(f"    Values with decimals: {dec_count}/{len(vals)} ({100*dec_count/len(vals):.0f}%)")

# Check formulas
print(f"\n  Formula checks:")
for col, name in [(HEIGHT_AVG, "J"), (BRANCH_AVG, "P"), (PODS_AVG, "V"),
                   (SEEDS_AVG, "AB"), (PODLEN_AVG, "AH"), (GRAIN_YIELD, "AJ")]:
    v = ws2.cell(row=3, column=col).value
    print(f"    Col {name} row 3: {v}")

print(f"\nDone! All changes applied to: {out}")
