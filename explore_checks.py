import openpyxl

wb = openpyxl.load_workbook(r"V:\temp\resume\Cowpea_data_final.xlsx", data_only=True)
ws = wb.active
print(f"Sheet: {ws.title}, Rows: {ws.max_row}, Cols: {ws.max_column}")

# Print headers (rows 1-2) for all columns up to AJ (col 36)
print("\n=== HEADERS (Rows 1-2) ===")
for col in range(1, min(ws.max_column + 1, 40)):
    letter = openpyxl.utils.get_column_letter(col)
    h1 = ws.cell(row=1, column=col).value
    h2 = ws.cell(row=2, column=col).value
    print(f"  Col {letter} ({col:2d}): Row1='{h1}' | Row2='{h2}'")

# First, check what's in column A and B for first few data rows
print("\n=== FIRST 5 DATA ROWS (row 3-7) - Cols A-F ===")
for row in range(3, 8):
    vals = []
    for col in range(1, 7):
        letter = openpyxl.utils.get_column_letter(col)
        v = ws.cell(row=row, column=col).value
        vals.append(f"{letter}={v}")
    print(f"  Row {row}: {' | '.join(vals)}")

# Now check each check group
check_groups = [
    ("Group 1", 20, 24),   # SL.No 18-22
    ("Group 2", 42, 46),   # SL.No 40-44
    ("Group 3", 64, 68),   # SL.No 62-66
    ("Group 4", 86, 90),   # SL.No 84-88
    ("Group 5", 107, 111), # SL.No 105-109
    ("Group 6", 130, 134), # SL.No 128-132
]

# Column mapping (will adjust after seeing headers)
# For now, let's dump cols A through AJ for each check row
print("\n=== CHECK VARIETIES - FULL DATA ===")
for group_name, start_row, end_row in check_groups:
    print(f"\n--- {group_name} (Excel rows {start_row}-{end_row}, expected SL.No {start_row-2}-{end_row-2}) ---")
    for row in range(start_row, end_row + 1):
        # First print SL.No and Genotype (cols A-D)
        sl = ws.cell(row=row, column=1).value
        vals_b = ws.cell(row=row, column=2).value
        vals_c = ws.cell(row=row, column=3).value
        vals_d = ws.cell(row=row, column=4).value
        print(f"  Row {row}: A(SL.No)={sl} | B={vals_b} | C={vals_c} | D={vals_d}")

        # Plant height: E-I reps, J avg
        ph = [ws.cell(row=row, column=c).value for c in range(5, 10)]
        ph_avg = ws.cell(row=row, column=10).value
        print(f"    Plant Height reps(E-I): {ph} | AVG(J): {ph_avg}")

        # Branches: K-O reps, P avg
        br = [ws.cell(row=row, column=c).value for c in range(11, 16)]
        br_avg = ws.cell(row=row, column=16).value
        print(f"    Branches reps(K-O): {br} | AVG(P): {br_avg}")

        # Pods: Q-U reps, V avg
        po = [ws.cell(row=row, column=c).value for c in range(17, 22)]
        po_avg = ws.cell(row=row, column=22).value
        print(f"    Pods reps(Q-U): {po} | AVG(V): {po_avg}")

        # Seeds per pod: W-AA reps, AB avg
        sp = [ws.cell(row=row, column=c).value for c in range(23, 28)]
        sp_avg = ws.cell(row=row, column=28).value
        print(f"    Seeds/pod reps(W-AA): {sp} | AVG(AB): {sp_avg}")

        # Pod length: AC-AG reps, AH avg
        pl = [ws.cell(row=row, column=c).value for c in range(29, 34)]
        pl_avg = ws.cell(row=row, column=34).value
        print(f"    Pod length reps(AC-AG): {pl} | AVG(AH): {pl_avg}")

        # Seed weight: AI, Grain yield: AJ
        sw = ws.cell(row=row, column=35).value
        gy = ws.cell(row=row, column=36).value
        print(f"    Seed weight(AI): {sw} | Grain yield(AJ): {gy}")

# Verify SL.No mapping
print("\n=== SL.No VERIFICATION ===")
print("Checking if Col A = SL.No and row 3 = SL.No 1:")
for row in [3, 4, 5, 20, 24, 42, 46, 64, 68, 86, 90, 107, 111, 130, 134]:
    sl = ws.cell(row=row, column=1).value
    geno = ws.cell(row=row, column=2).value
    print(f"  Excel row {row}: Col A = {sl}, Col B = {geno}")

# Check genotype consistency across groups
print("\n=== GENOTYPE CONSISTENCY CHECK ===")
all_group_names = []
for group_name, start_row, end_row in check_groups:
    names = []
    for row in range(start_row, end_row + 1):
        # Use whatever column has the genotype name
        name = ws.cell(row=row, column=2).value  # Try col B first
        if name is None:
            name = ws.cell(row=row, column=3).value  # Try col C
        names.append(name)
    all_group_names.append(names)
    print(f"  {group_name}: {names}")

# Check if all groups have same names
print("\nAre all 6 groups identical?")
first = all_group_names[0]
for i, names in enumerate(all_group_names[1:], 2):
    match = (names == first)
    print(f"  Group 1 vs Group {i}: {'MATCH' if match else 'DIFFER'}")
    if not match:
        for j, (a, b) in enumerate(zip(first, names)):
            if a != b:
                print(f"    Position {j+1}: '{a}' vs '{b}'")

wb.close()
