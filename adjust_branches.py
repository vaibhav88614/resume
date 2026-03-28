"""
Adjust branches per plant (cols K-O, avg P) to range 3-12,
correlated with pods per plant (cols Q-U), with randomness.
"""
import openpyxl
import random

random.seed(42)

src = r'V:\temp\resume\Cowpea_data_adjusted.xlsx'
wb = openpyxl.load_workbook(src)
ws = wb.active

# Columns: K=11, L=12, M=13, N=14, O=15 (branches reps 1-5)
#           P=16 (branches AVG formula)
#           Q=17, R=18, S=19, T=20, U=21 (pods reps 1-5)
BRANCH_COLS = [11, 12, 13, 14, 15]  # K-O
PODS_COLS   = [17, 18, 19, 20, 21]  # Q-U
AVG_COL = 16  # P

DATA_START = 3
DATA_END = 134  # rows 3-134 = 132 genotypes

TARGET_MIN = 3
TARGET_MAX = 12

# First pass: find pods range across all reps
all_pods = []
for row in range(DATA_START, DATA_END + 1):
    for col in PODS_COLS:
        v = ws.cell(row=row, column=col).value
        if v is not None:
            all_pods.append(v)

pods_min = min(all_pods)
pods_max = max(all_pods)
print(f"Pods per plant range: {pods_min} - {pods_max}")

# For each genotype row, compute branch values proportional to pod values
changed = 0

# Use percentile-based normalization for better spread
# Collect all per-row pod averages first
row_pod_avgs = []
for row in range(DATA_START, DATA_END + 1):
    pods_vals = []
    for col in PODS_COLS:
        v = ws.cell(row=row, column=col).value
        pods_vals.append(v if v is not None else 0)
    row_pod_avgs.append((row, sum(pods_vals) / len(pods_vals), pods_vals))

# Sort by pod average to assign rank-based branch centers
sorted_by_pods = sorted(row_pod_avgs, key=lambda x: x[1])
n = len(sorted_by_pods)

# Assign rank-based branch center: lowest pods -> ~4, highest pods -> ~11
for rank, (row, pods_avg, pods_vals) in enumerate(sorted_by_pods):
    # Rank-based norm (0 to 1)
    rank_norm = rank / max(n - 1, 1)
    
    # Branch center from rank: range 4 to 11 (leaving room for jitter)
    branch_center = 4 + rank_norm * 7
    
    # For each rep, add inter-rep variation based on pod deviations
    for i, col in enumerate(BRANCH_COLS):
        pod_val = pods_vals[i]
        
        # How much this rep deviates from the row's pod average
        if pods_avg > 0:
            rep_deviation = (pod_val - pods_avg) / pods_avg  # e.g. +0.15 or -0.10
        else:
            rep_deviation = 0
        
        # Scale deviation to ~±1.5 branches
        rep_offset = rep_deviation * 2.0
        
        # Random jitter ±1.5
        jitter = random.uniform(-1.5, 1.5)
        
        new_val = round(branch_center + rep_offset + jitter)
        
        # Clamp to target range
        new_val = max(TARGET_MIN, min(TARGET_MAX, new_val))
        
        ws.cell(row=row, column=col).value = new_val
        changed += 1
    
    # Preserve AVG formula in column P
    ws.cell(row=row, column=AVG_COL).value = f'=AVERAGE(K{row}:O{row})'

out_path = src
try:
    wb.save(out_path)
except PermissionError:
    out_path = src.replace('.xlsx', '_v2.xlsx')
    wb.save(out_path)
print(f"Saved to: {out_path}")
print(f"Updated {changed} branch rep values across {DATA_END - DATA_START + 1} rows")

# Verify
wb2 = openpyxl.load_workbook(out_path)
ws2 = wb2.active
branch_vals = []
for row in range(DATA_START, DATA_END + 1):
    for col in BRANCH_COLS:
        v = ws2.cell(row=row, column=col).value
        if v is not None:
            branch_vals.append(v)

print(f"\nVerification:")
print(f"  Branch values count: {len(branch_vals)}")
print(f"  Range: {min(branch_vals)} - {max(branch_vals)}")
print(f"  Mean: {sum(branch_vals)/len(branch_vals):.2f}")

# Distribution
from collections import Counter
dist = Counter(branch_vals)
print(f"  Distribution: {dict(sorted(dist.items()))}")

# Check correlation: print a few rows showing pods avg vs branches avg
print(f"\nSample rows (pods avg -> branches avg):")
for row in [3, 10, 20, 50, 80, 100, 130]:
    pods = [ws2.cell(row=row, column=c).value or 0 for c in PODS_COLS]
    branches = [ws2.cell(row=row, column=c).value or 0 for c in BRANCH_COLS]
    p_avg = sum(pods)/len(pods)
    b_avg = sum(branches)/len(branches)
    print(f"  Row {row}: pods_avg={p_avg:.1f}, branches_avg={b_avg:.1f}")
