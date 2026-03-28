import openpyxl
import statistics

files = {
    "Original": r"V:\temp\resume\Cowpea_data.xlsx",
    "Adjusted": r"V:\temp\resume\Cowpea_data_adjusted.xlsx",
    "Adjusted_v2": r"V:\temp\resume\Cowpea_data_adjusted_v2.xlsx",
}

DATA_ROWS = range(3, 135)  # rows 3-134 inclusive

def col_letter_to_idx(letter):
    """Convert column letter(s) to 1-based index."""
    result = 0
    for c in letter.upper():
        result = result * 26 + (ord(c) - ord('A') + 1)
    return result

# Column definitions
SEED_WEIGHT_COL = col_letter_to_idx("AI")  # 35
PODS_COLS = [col_letter_to_idx(c) for c in ["Q", "R", "S", "T", "U"]]  # 17-21
SEEDS_COLS = [col_letter_to_idx(c) for c in ["W", "X", "Y", "Z", "AA"]]  # 23-27
BRANCHES_COLS = [col_letter_to_idx(c) for c in ["K", "L", "M", "N", "O"]]  # 11-15
PLANT_HEIGHT_COLS = [col_letter_to_idx(c) for c in ["E", "F", "G", "H", "I"]]  # 5-9
POD_LENGTH_COLS = [col_letter_to_idx(c) for c in ["AC", "AD", "AE", "AF", "AG"]]  # 29-33
PLANT_HEIGHT_AVG_COL = col_letter_to_idx("J")  # 10
BRANCHES_AVG_COL = col_letter_to_idx("P")  # 16

for fname, fpath in files.items():
    print(f"\n{'='*70}")
    print(f"  FILE: {fname}")
    print(f"{'='*70}")
    
    try:
        # Open with data_only=False to detect formulas
        wb_formula = openpyxl.load_workbook(fpath, data_only=False)
        ws_f = wb_formula.active
        
        # Open with data_only=True to get computed values
        wb_data = openpyxl.load_workbook(fpath, data_only=True)
        ws = wb_data.active
    except FileNotFoundError:
        print(f"  *** FILE NOT FOUND ***")
        continue

    # --- Seed Weight (col AI) ---
    sw_vals = []
    for r in DATA_ROWS:
        v = ws.cell(row=r, column=SEED_WEIGHT_COL).value
        if v is not None and isinstance(v, (int, float)):
            sw_vals.append(v)
    print(f"\n  Seed Weight (AI, col 35): n={len(sw_vals)}")
    if sw_vals:
        print(f"    min={min(sw_vals):.4f}, max={max(sw_vals):.4f}, mean={statistics.mean(sw_vals):.4f}")

    # --- Pods per plant reps (Q-U) ---
    pods_vals = []
    for r in DATA_ROWS:
        for c in PODS_COLS:
            v = ws.cell(row=r, column=c).value
            if v is not None and isinstance(v, (int, float)):
                pods_vals.append(v)
    print(f"\n  Pods per Plant reps (Q-U, cols 17-21): n={len(pods_vals)}")
    if pods_vals:
        print(f"    min={min(pods_vals):.4f}, max={max(pods_vals):.4f}, mean={statistics.mean(pods_vals):.4f}")

    # --- Seeds per pod reps (W-AA) ---
    seeds_vals = []
    for r in DATA_ROWS:
        for c in SEEDS_COLS:
            v = ws.cell(row=r, column=c).value
            if v is not None and isinstance(v, (int, float)):
                seeds_vals.append(v)
    print(f"\n  Seeds per Pod reps (W-AA, cols 23-27): n={len(seeds_vals)}")
    if seeds_vals:
        print(f"    min={min(seeds_vals):.4f}, max={max(seeds_vals):.4f}, range={min(seeds_vals):.4f}-{max(seeds_vals):.4f}, mean={statistics.mean(seeds_vals):.4f}")

    # --- Branches per plant reps (K-O) ---
    br_vals = []
    for r in DATA_ROWS:
        for c in BRANCHES_COLS:
            v = ws.cell(row=r, column=c).value
            if v is not None and isinstance(v, (int, float)):
                br_vals.append(v)
    print(f"\n  Branches per Plant reps (K-O, cols 11-15): n={len(br_vals)}")
    if br_vals:
        print(f"    min={min(br_vals):.4f}, max={max(br_vals):.4f}, mean={statistics.mean(br_vals):.4f}")

    # --- Plant Height reps (E-I): integer vs decimal check ---
    ph_vals = []
    ph_float_count = 0
    for r in DATA_ROWS:
        for c in PLANT_HEIGHT_COLS:
            v = ws.cell(row=r, column=c).value
            if v is not None and isinstance(v, (int, float)):
                ph_vals.append(v)
                if isinstance(v, float) and v != int(v):
                    ph_float_count += 1
    print(f"\n  Plant Height reps (E-I, cols 5-9): n={len(ph_vals)}")
    if ph_vals:
        print(f"    min={min(ph_vals):.4f}, max={max(ph_vals):.4f}")
        print(f"    Values with non-zero decimal part: {ph_float_count} / {len(ph_vals)}")
        if ph_float_count > 0:
            # Show a few examples
            examples = []
            for r in DATA_ROWS:
                for c in PLANT_HEIGHT_COLS:
                    v = ws.cell(row=r, column=c).value
                    if v is not None and isinstance(v, float) and v != int(v):
                        examples.append(v)
                        if len(examples) >= 5:
                            break
                if len(examples) >= 5:
                    break
            print(f"    Examples of decimal values: {examples}")

    # --- Pod Length reps (AC-AG): integer vs decimal check ---
    pl_vals = []
    pl_float_count = 0
    for r in DATA_ROWS:
        for c in POD_LENGTH_COLS:
            v = ws.cell(row=r, column=c).value
            if v is not None and isinstance(v, (int, float)):
                pl_vals.append(v)
                if isinstance(v, float) and v != int(v):
                    pl_float_count += 1
    print(f"\n  Pod Length reps (AC-AG, cols 29-33): n={len(pl_vals)}")
    if pl_vals:
        print(f"    min={min(pl_vals):.4f}, max={max(pl_vals):.4f}")
        print(f"    Values with non-zero decimal part: {pl_float_count} / {len(pl_vals)}")
        if pl_float_count > 0:
            examples = []
            for r in DATA_ROWS:
                for c in POD_LENGTH_COLS:
                    v = ws.cell(row=r, column=c).value
                    if v is not None and isinstance(v, float) and v != int(v):
                        examples.append(v)
                        if len(examples) >= 5:
                            break
                if len(examples) >= 5:
                    break
            print(f"    Examples of decimal values: {examples}")

    # --- Check if col J (plant height AVG) and P (branches AVG) are formulas ---
    print(f"\n  Column J (Plant Height AVG) - formula check:")
    formula_count_j = 0
    hardcoded_count_j = 0
    for r in DATA_ROWS:
        cell = ws_f.cell(row=r, column=PLANT_HEIGHT_AVG_COL)
        if cell.value is not None:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_count_j += 1
            else:
                hardcoded_count_j += 1
    print(f"    Formulas: {formula_count_j}, Hardcoded: {hardcoded_count_j}")
    # Show example
    ex_cell = ws_f.cell(row=3, column=PLANT_HEIGHT_AVG_COL)
    print(f"    Row 3 value (raw): {ex_cell.value}")

    print(f"\n  Column P (Branches AVG) - formula check:")
    formula_count_p = 0
    hardcoded_count_p = 0
    for r in DATA_ROWS:
        cell = ws_f.cell(row=r, column=BRANCHES_AVG_COL)
        if cell.value is not None:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_count_p += 1
            else:
                hardcoded_count_p += 1
    print(f"    Formulas: {formula_count_p}, Hardcoded: {hardcoded_count_p}")
    ex_cell = ws_f.cell(row=3, column=BRANCHES_AVG_COL)
    print(f"    Row 3 value (raw): {ex_cell.value}")
    
    wb_formula.close()
    wb_data.close()

print("\n\nDone.")
