import openpyxl
from statistics import mean

wb = openpyxl.load_workbook(r"V:\temp\resume\Cowpea_data_final.xlsx", data_only=True)
ws = wb.active

# Check row ranges per group (Excel 1-based rows)
group_rows = {
    1: range(20, 25),   # rows 20-24
    2: range(42, 47),   # rows 42-46
    3: range(64, 69),   # rows 64-68
    4: range(86, 91),   # rows 86-90
    5: range(107, 112), # rows 107-111
    6: range(130, 135), # rows 130-134
}

varieties = [
    "C1 C-152",
    "C2 KBC-2",
    "C3 KBC-9",
    "C4 IT-38956-1",
    "C5 LOCAL VARIETY",
]

# Parameter definitions: name -> (col_start, col_end) 1-based, inclusive
# For multi-rep params: 5 reps each
params_multi = {
    "Plant height":  (5, 9),    # E-I
    "Branches":      (11, 15),  # K-O
    "Pods/plant":    (17, 21),  # Q-U
    "Seeds/pod":     (23, 27),  # W-AA
    "Pod length":    (29, 33),  # AC-AG
}
params_single = {
    "Seed weight":      35,  # AI
}
params_once = {
    "Days to flowering": 3,  # C
    "Days to maturity":  4,  # D
}

def safe_float(v):
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None

# First, print what's actually in these cells for verification
print("=" * 80)
print("VERIFICATION: Variety names in column B (col 2) for each group")
print("=" * 80)
for g, rows in group_rows.items():
    print(f"\nGroup {g}:")
    for r in rows:
        val_a = ws.cell(row=r, column=1).value  # SL.No
        val_b = ws.cell(row=r, column=2).value  # Variety name
        print(f"  Row {r}: SL.No={val_a}, Name={val_b}")

print("\n" + "=" * 120)
print("DETAILED CHECK VARIETY DATA ACROSS 6 GROUPS")
print("=" * 120)

# Collect all data organized by variety
all_data = {}  # variety_idx -> param_name -> list of values (per group)

for vi in range(5):  # 5 varieties
    vname = varieties[vi]
    all_data[vi] = {}

    # Multi-rep parameters
    for pname, (cs, ce) in params_multi.items():
        group_vals = []
        for g in range(1, 7):
            rows = list(group_rows[g])
            r = rows[vi]  # row for this variety in this group
            reps = []
            for c in range(cs, ce + 1):
                reps.append(safe_float(ws.cell(row=r, column=c).value))
            group_vals.append(reps)
        all_data[vi][pname] = group_vals

    # Single-value parameters (per group)
    for pname, col in params_single.items():
        group_vals = []
        for g in range(1, 7):
            rows = list(group_rows[g])
            r = rows[vi]
            group_vals.append(safe_float(ws.cell(row=r, column=col).value))
        all_data[vi][pname] = group_vals

    # Once-only parameters (take from group 1)
    for pname, col in params_once.items():
        rows = list(group_rows[1])
        r = rows[vi]
        val = safe_float(ws.cell(row=r, column=col).value)
        # Also check all groups for these
        gvals = []
        for g in range(1, 7):
            rows_g = list(group_rows[g])
            rg = rows_g[vi]
            gvals.append(safe_float(ws.cell(row=rg, column=col).value))
        all_data[vi][pname] = gvals

# Print structured output
for vi in range(5):
    vname = varieties[vi]
    print(f"\n{'#' * 120}")
    print(f"### VARIETY: {vname}")
    print(f"{'#' * 120}")

    # Days to flowering / maturity
    for pname in ["Days to flowering", "Days to maturity"]:
        gvals = all_data[vi][pname]
        nums = [v for v in gvals if v is not None]
        m = mean(nums) if nums else None
        rng = max(nums) - min(nums) if len(nums) >= 2 else 0
        print(f"\n  {pname}:")
        print(f"    Groups 1-6: {gvals}")
        print(f"    Mean={m:.2f}, Range={rng:.2f}" if m else f"    Mean=N/A, Range=N/A")

    # Multi-rep parameters
    for pname in ["Plant height", "Branches", "Pods/plant", "Seeds/pod", "Pod length"]:
        gvals = all_data[vi][pname]  # list of 6 lists (each 5 reps)
        print(f"\n  {pname}:")
        group_means = []
        for g_idx, reps in enumerate(gvals):
            nums = [v for v in reps if v is not None]
            gm = mean(nums) if nums else None
            group_means.append(gm)
            print(f"    Group {g_idx+1}: reps={reps}  mean={gm:.2f}" if gm else f"    Group {g_idx+1}: reps={reps}  mean=N/A")
        valid_means = [v for v in group_means if v is not None]
        if valid_means:
            overall = mean(valid_means)
            rng = max(valid_means) - min(valid_means)
            print(f"    >>> Overall mean of group-means = {overall:.2f}, Range across groups = {rng:.2f}")

    # Seed weight (single value per group)
    pname = "Seed weight"
    gvals = all_data[vi][pname]
    nums = [v for v in gvals if v is not None]
    m = mean(nums) if nums else None
    rng = max(nums) - min(nums) if len(nums) >= 2 else 0
    print(f"\n  {pname}:")
    print(f"    Groups 1-6: {gvals}")
    print(f"    Mean={m:.2f}, Range={rng:.2f}" if m else f"    Mean=N/A, Range=N/A")

# Summary table
print("\n\n" + "=" * 120)
print("SUMMARY: MEAN AND RANGE OF GROUP-MEANS FOR EACH VARIETY × PARAMETER")
print("=" * 120)
header = f"{'Variety':<22} {'Parameter':<18} {'G1':>8} {'G2':>8} {'G3':>8} {'G4':>8} {'G5':>8} {'G6':>8} {'Mean':>8} {'Range':>8}"
print(header)
print("-" * len(header))

for vi in range(5):
    vname = varieties[vi]
    for pname in ["Days to flowering", "Days to maturity", "Plant height", "Branches",
                  "Pods/plant", "Seeds/pod", "Pod length", "Seed weight"]:
        gvals = all_data[vi][pname]
        if pname in params_multi:
            # Compute group means from reps
            display_vals = []
            for reps in gvals:
                nums = [v for v in reps if v is not None]
                display_vals.append(mean(nums) if nums else None)
        else:
            display_vals = gvals

        valid = [v for v in display_vals if v is not None]
        m = mean(valid) if valid else None
        rng = max(valid) - min(valid) if len(valid) >= 2 else None

        vals_str = ""
        for dv in display_vals:
            vals_str += f"{dv:8.2f}" if dv is not None else f"{'N/A':>8}"
        m_str = f"{m:8.2f}" if m is not None else f"{'N/A':>8}"
        r_str = f"{rng:8.2f}" if rng is not None else f"{'N/A':>8}"

        print(f"{vname:<22} {pname:<18} {vals_str} {m_str} {r_str}")
    print()

wb.close()
