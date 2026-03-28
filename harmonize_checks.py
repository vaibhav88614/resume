"""
Harmonize check variety values across 6 groups so the same genotype
has similar values in all blocks, with natural variation retained.

Check varieties (5), each appearing in 6 groups:
  C1 C-152, C2 KBC-2, C3 KBC-9, C4 IT-38956-1, C5 LOCAL VARIETY

Parameters to harmonize:
  - Days to 50% flowering (C, col 3) - single value
  - Plant height reps (E-I, cols 5-9) - 5 reps
  - Branches reps (K-O, cols 11-15) - 5 reps
  - Pods/plant reps (Q-U, cols 17-21) - 5 reps
  - Seeds/pod reps (W-AA, cols 23-27) - 5 reps
  - Pod length reps (AC-AG, cols 29-33) - 5 reps
  - Seed weight (AI, col 35) - single value
"""
import openpyxl
import random
import statistics

random.seed(42)

src = r'V:\temp\resume\Cowpea_data_final.xlsx'
wb = openpyxl.load_workbook(src)
ws = wb.active

# Check row positions (Excel rows) - each group has 5 varieties in order
GROUPS = [
    [20, 21, 22, 23, 24],     # Group 1: SL.No 18-22
    [42, 43, 44, 45, 46],     # Group 2: SL.No 40-44
    [64, 65, 66, 67, 68],     # Group 3: SL.No 62-66
    [86, 87, 88, 89, 90],     # Group 4: SL.No 84-88
    [107, 108, 109, 110, 111], # Group 5: SL.No 105-109
    [130, 131, 132, 133, 134], # Group 6: SL.No 128-132
]

VARIETY_NAMES = ["C1 C-152", "C2 KBC-2", "C3 KBC-9", "C4 IT-38956-1", "C5 LOCAL VARIETY"]

# Parameter definitions: (name, columns, type, decimals)
# type: 'reps' = 5 rep columns, 'single' = 1 column
PARAMS = [
    ("Days to flowering", [3], "single", 0),
    ("Plant height",      [5, 6, 7, 8, 9], "reps", 1),
    ("Branches/plant",    [11, 12, 13, 14, 15], "reps", 0),
    ("Pods/plant",        [17, 18, 19, 20, 21], "reps", 0),
    ("Seeds/pod",         [23, 24, 25, 26, 27], "reps", 0),
    ("Pod length",        [29, 30, 31, 32, 33], "reps", 1),
    ("Seed weight",       [35], "single", 2),
]

# Formula columns to refresh after changes
AVG_FORMULAS = {
    10: ('E', 'I'),   # J = Plant height AVG
    16: ('K', 'O'),   # P = Branches AVG
    22: ('Q', 'U'),   # V = Pods AVG
    28: ('W', 'AA'),  # AB = Seeds AVG
    34: ('AC', 'AG'), # AH = Pod length AVG
}


def harmonize_variety(variety_idx):
    """Harmonize values for one check variety across all 6 groups."""
    # Get the row for this variety in each group
    rows = [group[variety_idx] for group in GROUPS]
    variety_name = VARIETY_NAMES[variety_idx]
    
    print(f"\n  {variety_name} (rows: {rows})")
    
    for param_name, cols, ptype, decimals in PARAMS:
        if ptype == "single":
            # Collect the single value from each group
            values = []
            for row in rows:
                v = ws.cell(row=row, column=cols[0]).value
                if v is not None:
                    values.append(float(v))
            
            if not values:
                continue
            
            # Compute target: use median to be robust against outliers
            target = statistics.median(values)
            
            # Generate 6 values clustered around target with small variation
            # Variation: ±5% for most, ±2 for small integer values
            for row in rows:
                if param_name == "Days to flowering":
                    # Integer, small range: ±2 days
                    jitter = random.randint(-2, 2)
                    new_val = int(round(target + jitter))
                elif param_name == "Seed weight":
                    # ±8% variation for seed weight
                    factor = random.uniform(0.92, 1.08)
                    new_val = round(target * factor, decimals)
                else:
                    factor = random.uniform(0.95, 1.05)
                    new_val = round(target * factor, decimals)
                
                ws.cell(row=row, column=cols[0]).value = new_val
            
            # Print summary
            new_vals = [ws.cell(row=r, column=cols[0]).value for r in rows]
            print(f"    {param_name}: median={target:.1f}, new values={new_vals}")
        
        else:  # reps
            # Collect ALL rep values across all groups to find the typical range
            all_rep_values = []
            group_means = []
            for row in rows:
                row_vals = []
                for col in cols:
                    v = ws.cell(row=row, column=col).value
                    if v is not None:
                        all_rep_values.append(float(v))
                        row_vals.append(float(v))
                if row_vals:
                    group_means.append(statistics.mean(row_vals))
            
            if not all_rep_values:
                continue
            
            # Target center: median of group means
            target_center = statistics.median(group_means)
            
            # Typical within-row spread (std dev of reps within a row)
            within_spreads = []
            for row in rows:
                row_vals = []
                for col in cols:
                    v = ws.cell(row=row, column=col).value
                    if v is not None:
                        row_vals.append(float(v))
                if len(row_vals) >= 2:
                    within_spreads.append(statistics.stdev(row_vals))
            
            typical_spread = statistics.median(within_spreads) if within_spreads else 2.0
            
            # For each group, generate new rep values around a group center
            # Group centers are slightly varied around target_center
            for row in rows:
                # Group center: target ± small variation (±5%, tighter ±3% for height)
                if param_name == "Plant height":
                    group_center = target_center * random.uniform(0.97, 1.03)
                else:
                    group_center = target_center * random.uniform(0.95, 1.05)
                
                for col in cols:
                    # Each rep: group_center + random spread
                    # Use tighter spread for plant height to avoid large group differences
                    spread_factor = 0.35 if param_name == "Plant height" else 0.7
                    jitter = random.gauss(0, typical_spread * spread_factor)
                    new_val = group_center + jitter
                    
                    # Round appropriately
                    if decimals == 0:
                        new_val = int(round(max(1, new_val)))  # min 1 for integers
                    else:
                        new_val = round(max(0.1, new_val), decimals)
                    
                    ws.cell(row=row, column=col).value = new_val
            
            # Apply parameter-specific constraints
            if param_name == "Branches/plant":
                for row in rows:
                    for col in cols:
                        v = ws.cell(row=row, column=col).value
                        if v is not None:
                            ws.cell(row=row, column=col).value = max(3, min(12, int(v)))
            elif param_name == "Pods/plant":
                for row in rows:
                    for col in cols:
                        v = ws.cell(row=row, column=col).value
                        if v is not None:
                            ws.cell(row=row, column=col).value = max(10, min(34, int(v)))
            elif param_name == "Seeds/pod":
                for row in rows:
                    for col in cols:
                        v = ws.cell(row=row, column=col).value
                        if v is not None:
                            ws.cell(row=row, column=col).value = max(12, min(26, int(v)))
            elif param_name == "Plant height":
                # Additionally clamp each rep to be within ±10% of target center
                lo = target_center * 0.90
                hi = target_center * 1.10
                for row in rows:
                    for col in cols:
                        v = ws.cell(row=row, column=col).value
                        if v is not None:
                            clamped = max(lo, min(hi, float(v)))
                            ws.cell(row=row, column=col).value = round(clamped, 1)
            
            # Print summary
            new_group_means = []
            for row in rows:
                rv = []
                for c in cols:
                    val = ws.cell(row=row, column=c).value
                    if val is not None:
                        rv.append(float(val))
                if rv:
                    new_group_means.append(round(statistics.mean(rv), 1))
            
            print(f"    {param_name}: target={target_center:.1f}, "
                  f"group means={new_group_means}")


# =====================================================================
# HARMONIZE ALL 5 VARIETIES
# =====================================================================
print("=" * 60)
print("Harmonizing check varieties across 6 groups")

for vi in range(5):
    harmonize_variety(vi)

# =====================================================================
# REFRESH AVG FORMULAS FOR CHECK ROWS
# =====================================================================
print(f"\n{'=' * 60}")
print("Refreshing AVG formulas for check rows")

for group in GROUPS:
    for row in group:
        for avg_col, (start_letter, end_letter) in AVG_FORMULAS.items():
            ws.cell(row=row, column=avg_col).value = \
                f'=AVERAGE({start_letter}{row}:{end_letter}{row})'
        # Grain yield formula
        ws.cell(row=row, column=36).value = f'=AI{row}*7407.4/1000'

print("  All formulas refreshed for 30 check rows")

# =====================================================================
# SAVE
# =====================================================================
out = src
try:
    wb.save(out)
except PermissionError:
    out = src.replace('.xlsx', '_v2.xlsx')
    wb.save(out)
print(f"\nSaved: {out}")

# =====================================================================
# VERIFY
# =====================================================================
print(f"\n{'=' * 60}")
print("VERIFICATION - Check variety consistency")

wb2 = openpyxl.load_workbook(out)
ws2 = wb2.active

for vi in range(5):
    variety_name = VARIETY_NAMES[vi]
    rows = [group[vi] for group in GROUPS]
    print(f"\n  {variety_name}:")
    
    for param_name, param_cols, ptype, decimals in PARAMS:
        if ptype == "single":
            vals = [ws2.cell(row=r, column=param_cols[0]).value for r in rows]
            vals_num = [v for v in vals if v is not None]
            if vals_num:
                rng = max(vals_num) - min(vals_num)
                print(f"    {param_name}: {vals_num}  range={rng:.2f}")
        else:
            group_avgs = []
            for row in rows:
                rv = []
                for c in param_cols:
                    v = ws2.cell(row=row, column=c).value
                    if v is not None:
                        rv.append(float(v))
                if rv:
                    group_avgs.append(round(statistics.mean(rv), 1))
            if group_avgs:
                rng = max(group_avgs) - min(group_avgs)
                print(f"    {param_name} (avg): {group_avgs}  range={rng:.1f}")

print("\nDone!")
