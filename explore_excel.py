import openpyxl
import os
from openpyxl.utils import get_column_letter

# Try original first, then v2
for fname in ['Cowpea_data_adjusted.xlsx', 'Cowpea_data_adjusted_v2.xlsx']:
    fpath = os.path.join(r'V:\temp\resume', fname)
    if os.path.exists(fpath):
        try:
            wb = openpyxl.load_workbook(fpath, data_only=True)
            print(f'Opened: {fname}')
            break
        except Exception as e:
            print(f'Cannot open {fname}: {e}')
            continue
else:
    print('Available xlsx files:')
    for f in os.listdir(r'V:\temp\resume'):
        if f.endswith('.xlsx'):
            print(f'  {f}')
    exit()

ws = wb.active
print(f'Sheet: {ws.title}')
print(f'Dimensions: {ws.dimensions}')
print(f'Max row: {ws.max_row}, Max col: {ws.max_column}')
print()

# 1. ALL column headers (rows 1-2)
print('=' * 120)
print('ALL COLUMN HEADERS (Row 1 and Row 2)')
print('=' * 120)
for col in range(1, ws.max_column + 1):
    letter = get_column_letter(col)
    r1 = ws.cell(row=1, column=col).value
    r2 = ws.cell(row=2, column=col).value
    print(f'  Col {letter:>3} (#{col:>3}): Row1={repr(r1):<55} Row2={repr(r2)}')

print()

# 2. Merged cells in header rows
print('=' * 80)
print('MERGED CELLS (affecting rows 1-2)')
print('=' * 80)
for mc in ws.merged_cells.ranges:
    if mc.min_row <= 2:
        print(f'  {mc}')

print()

# 3. Search for plant height and pod length
print('=' * 80)
print('KEYWORD SEARCH IN ALL HEADERS')
print('=' * 80)
keywords = ['height', 'plant h', 'ht', 'pod length', 'length', 'branch', 'pods', 'seed']
for col in range(1, ws.max_column + 1):
    letter = get_column_letter(col)
    r1 = str(ws.cell(row=1, column=col).value or '').lower()
    r2 = str(ws.cell(row=2, column=col).value or '').lower()
    combined = r1 + ' ' + r2
    for kw in keywords:
        if kw in combined:
            print(f'  [{kw:>12}] Col {letter} (#{col}): Row1={ws.cell(row=1, column=col).value!r}, Row2={ws.cell(row=2, column=col).value!r}')
            break

print()

# 4. Detailed analysis of columns E through AJ (cols 5-36)
print('=' * 80)
print('DETAILED COLUMN ANALYSIS: E through end')
print('=' * 80)
for col in range(5, ws.max_column + 1):
    letter = get_column_letter(col)
    r1 = ws.cell(row=1, column=col).value
    r2 = ws.cell(row=2, column=col).value
    
    # Collect all data values
    all_vals = []
    for r in range(3, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if v is not None and isinstance(v, (int, float)):
            all_vals.append(v)
    
    sample = [ws.cell(row=r, column=col).value for r in range(3, 11)]
    
    int_count = sum(1 for v in all_vals if isinstance(v, int))
    float_count = sum(1 for v in all_vals if isinstance(v, float))
    whole_floats = sum(1 for v in all_vals if isinstance(v, float) and v == int(v))
    
    # Check if column has formulas (load without data_only)
    stats_str = ''
    if all_vals:
        stats_str = f'min={min(all_vals)}, max={max(all_vals)}, mean={sum(all_vals)/len(all_vals):.2f}, count={len(all_vals)}'
    
    is_avg = r2 and 'avg' in str(r2).lower()
    col_type = 'AVG' if is_avg else 'REP'
    
    print(f'  Col {letter:>2} (#{col:>2}) [{col_type}]: H1={repr(r1):<40} H2={repr(r2):<15}')
    print(f'         Sample(3-10): {sample}')
    print(f'         Types: {int_count} ints, {float_count} floats ({whole_floats} whole-number floats)')
    if stats_str:
        print(f'         Stats: {stats_str}')
    print()

# 5. Also check columns A-D
print('=' * 80)
print('COLUMNS A-D (identifiers)')
print('=' * 80)
for col in range(1, 5):
    letter = get_column_letter(col)
    r1 = ws.cell(row=1, column=col).value
    r2 = ws.cell(row=2, column=col).value
    sample = [ws.cell(row=r, column=col).value for r in range(3, 8)]
    print(f'  Col {letter}: H1={repr(r1)}, H2={repr(r2)}, sample={sample}')

# 6. Check formulas version
print()
print('=' * 80)
print('FORMULA CHECK (non data_only)')
print('=' * 80)
wb2 = openpyxl.load_workbook(fpath, data_only=False)
ws2 = wb2.active
# Check AVG columns: look for columns where row2 says AVG or similar
for col in range(5, ws2.max_column + 1):
    letter = get_column_letter(col)
    r2 = ws2.cell(row=2, column=col).value
    cell_val = ws2.cell(row=3, column=col).value
    if isinstance(cell_val, str) and cell_val.startswith('='):
        print(f'  Col {letter} (#{col}): H2={repr(r2)}, Formula={cell_val}')
wb2.close()

wb.close()
print('\nDone!')
